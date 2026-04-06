"""
Excel Report Generator — v4
- Delayed → Yellow, Missing POD → Red, High Travel → Orange
- Performance Score colour banding (green/amber/red)
- Store Insights sheet
- Daily Insights sheet (efficiency, working hours, score)
- Route Summary: Stores per Hour, Efficiency %, Avg Perf Score
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import io

# ── Colour palette ────────────────────────────────────────────────────────────
C_HDR_BG  = '1A3C5E'
C_HDR_FG  = 'FFFFFF'
C_ALT     = 'EBF3FB'
C_OK      = 'D4EDDA'   # green
C_WARN    = 'FFF3CD'   # yellow  — Delayed
C_ORANGE  = 'FFE0B2'   # orange  — High Travel
C_ERR     = 'F8D7DA'   # red     — Missing POD
C_BLUE    = 'DBEAFE'   # light blue — informational

# Score banding
C_SCORE_HI  = 'D4EDDA'   # ≥80  green
C_SCORE_MID = 'FFF3CD'   # 50–79 amber
C_SCORE_LO  = 'F8D7DA'   # <50  red

_thin  = Side(style='thin', color='C0C0C0')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

STATUS_COLORS = {
    'OK':          C_OK,
    'Complete':    C_OK,
    'Delayed':     C_WARN,     # yellow
    'High Travel': C_ORANGE,   # orange
    'Missing POD': C_ERR,      # red
    'Not Ended':   C_ERR,
}


# ── Primitives ────────────────────────────────────────────────────────────────

def _hdr(ws, headers: list, bg: str = C_HDR_BG):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color=C_HDR_FG, name='Arial', size=10)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = BORDER
    ws.row_dimensions[1].height = 28


def _row(ws, row_num: int, values: list, alt: bool = False):
    bg = C_ALT if alt else 'FFFFFF'
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=v)
        c.font      = Font(name='Arial', size=9)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = BORDER


def _autowidth(ws, mn: int = 9, mx: int = 38):
    for cc in ws.columns:
        w = max(len(str(c.value or '')) for c in cc)
        ws.column_dimensions[get_column_letter(cc[0].column)].width = min(max(w + 2, mn), mx)


def _cell_color(ws, row_num: int, headers: list, col_name: str, color: str,
                bold: bool = False):
    if col_name in headers:
        c = ws.cell(row_num, headers.index(col_name) + 1)
        c.fill = PatternFill('solid', fgColor=color)
        if bold:
            c.font = Font(name='Arial', size=9, bold=True)


def _write_sheet(wb, sheet_name: str, rows: list,
                 status_col: str = 'Status',
                 hdr_bg: str = C_HDR_BG,
                 extra_color_fn=None) -> None:
    ws = wb.create_sheet(sheet_name)
    if not rows:
        ws.cell(1, 1, 'No data.')
        return
    headers = list(rows[0].keys())
    _hdr(ws, headers, bg=hdr_bg)
    ws.freeze_panes = 'A2'

    for i, r in enumerate(rows):
        rn = i + 2
        _row(ws, rn, list(r.values()), alt=(i % 2 == 0))

        # Status colour
        if status_col in headers:
            sv = r.get(status_col, '')
            if sv in STATUS_COLORS:
                c = ws.cell(rn, headers.index(status_col) + 1)
                c.fill = PatternFill('solid', fgColor=STATUS_COLORS[sv])
                c.font = Font(name='Arial', size=9, bold=True)

        if extra_color_fn:
            extra_color_fn(ws, rn, headers, r)

    _autowidth(ws)


# ── Per-sheet colour callbacks ────────────────────────────────────────────────

def _color_details(ws, rn, headers, r):
    status = r.get('Status', '')

    # Travel Time: >60 → red, >30 → orange
    if 'Travel Time (mins)' in headers:
        ci = headers.index('Travel Time (mins)') + 1
        try:
            v = float(r['Travel Time (mins)'])
            bg = C_ERR if v > 60 else (C_ORANGE if v > 30 else None)
            if bg:
                ws.cell(rn, ci).fill = PatternFill('solid', fgColor=bg)
        except (ValueError, TypeError):
            pass

    # Store Time: >30 → yellow
    if 'Store Time (mins)' in headers:
        ci2 = headers.index('Store Time (mins)') + 1
        try:
            v2 = float(r['Store Time (mins)'])
            if v2 > 30:
                ws.cell(rn, ci2).fill = PatternFill('solid', fgColor=C_WARN)
        except (ValueError, TypeError):
            pass

    # Performance Score colour banding
    if 'Perf Score' in headers:
        ci3 = headers.index('Perf Score') + 1
        try:
            sc = int(r.get('Perf Score', 100))
            bg3 = C_SCORE_HI if sc >= 80 else (C_SCORE_MID if sc >= 50 else C_SCORE_LO)
            c3 = ws.cell(rn, ci3)
            c3.fill = PatternFill('solid', fgColor=bg3)
            c3.font = Font(name='Arial', size=9, bold=True)
        except (ValueError, TypeError):
            pass


def _color_routes(ws, rn, headers, r):
    # Avg Travel >45 → orange
    if 'Avg Travel Time (mins)' in headers:
        ci = headers.index('Avg Travel Time (mins)') + 1
        try:
            v = float(r.get('Avg Travel Time (mins)') or 0)
            if v > 45:
                ws.cell(rn, ci).fill = PatternFill('solid', fgColor=C_ORANGE)
        except (ValueError, TypeError):
            pass

    # Stores per Hour: >2 → green, <1 → amber
    if 'Stores per Hour' in headers:
        ci2 = headers.index('Stores per Hour') + 1
        try:
            v2 = float(r.get('Stores per Hour') or 0)
            bg2 = C_OK if v2 >= 2 else (C_WARN if v2 < 1 else None)
            if bg2:
                ws.cell(rn, ci2).fill = PatternFill('solid', fgColor=bg2)
        except (ValueError, TypeError):
            pass

    # Efficiency %: <70 → amber, <50 → red
    if 'Efficiency %' in headers:
        ci3 = headers.index('Efficiency %') + 1
        try:
            v3 = float(r.get('Efficiency %') or 100)
            bg3 = C_ERR if v3 < 50 else (C_WARN if v3 < 70 else C_OK)
            ws.cell(rn, ci3).fill = PatternFill('solid', fgColor=bg3)
        except (ValueError, TypeError):
            pass

    # Avg Perf Score banding
    if 'Avg Perf Score' in headers:
        ci4 = headers.index('Avg Perf Score') + 1
        try:
            sc = float(r.get('Avg Perf Score') or 100)
            bg4 = C_SCORE_HI if sc >= 80 else (C_SCORE_MID if sc >= 50 else C_SCORE_LO)
            ws.cell(rn, ci4).fill = PatternFill('solid', fgColor=bg4)
        except (ValueError, TypeError):
            pass


def _color_summary(ws, rn, headers, r):
    # Net Working Time < 300 → amber
    if 'Net Working Time (mins)' in headers:
        ci = headers.index('Net Working Time (mins)') + 1
        try:
            v = float(r.get('Net Working Time (mins)') or 0)
            if v < 300:
                ws.cell(rn, ci).fill = PatternFill('solid', fgColor=C_WARN)
        except (ValueError, TypeError):
            pass

    # Efficiency %
    if 'Efficiency %' in headers:
        ci2 = headers.index('Efficiency %') + 1
        try:
            v2 = float(r.get('Efficiency %') or 100)
            bg2 = C_ERR if v2 < 50 else (C_WARN if v2 < 70 else C_OK)
            ws.cell(rn, ci2).fill = PatternFill('solid', fgColor=bg2)
        except (ValueError, TypeError):
            pass

    # Avg Perf Score banding
    if 'Avg Perf Score' in headers:
        ci3 = headers.index('Avg Perf Score') + 1
        try:
            sc = float(r.get('Avg Perf Score') or 100)
            bg3 = C_SCORE_HI if sc >= 80 else (C_SCORE_MID if sc >= 50 else C_SCORE_LO)
            ws.cell(rn, ci3).fill = PatternFill('solid', fgColor=bg3)
        except (ValueError, TypeError):
            pass

    # Missing POD count — any > 0 → red
    if 'Missing POD' in headers:
        ci4 = headers.index('Missing POD') + 1
        try:
            if int(r.get('Missing POD') or 0) > 0:
                ws.cell(rn, ci4).fill = PatternFill('solid', fgColor=C_ERR)
        except (ValueError, TypeError):
            pass

    # Delayed count — any > 0 → yellow
    if 'Delayed' in headers:
        ci5 = headers.index('Delayed') + 1
        try:
            if int(r.get('Delayed') or 0) > 0:
                ws.cell(rn, ci5).fill = PatternFill('solid', fgColor=C_WARN)
        except (ValueError, TypeError):
            pass


def _color_exc(ws, rn, headers, r):
    KIND = {
        'Missing POD': C_ERR, 'Route Not Ended': C_ERR,
        'POD Without Store': C_ERR, 'Route End Without Start': C_ERR,
        'Long Delay': C_WARN, 'No Activity Gap': C_WARN,
        'High Travel Time': C_ORANGE, 'Break End Without Start': C_WARN,
    }
    bg = KIND.get(r.get('Exception Type', ''), C_WARN)
    for col in range(1, len(headers) + 1):
        ws.cell(rn, col).fill = PatternFill('solid', fgColor=bg)


def _color_store_insights(ws, rn, headers, r):
    # High visit count → blue highlight
    if 'Total Visits' in headers:
        ci = headers.index('Total Visits') + 1
        try:
            v = int(r.get('Total Visits') or 0)
            if v >= 5:
                ws.cell(rn, ci).fill = PatternFill('solid', fgColor=C_BLUE)
        except (ValueError, TypeError):
            pass
    # Missing POD > 0 → red
    if 'Missing POD Count' in headers:
        ci2 = headers.index('Missing POD Count') + 1
        try:
            if int(r.get('Missing POD Count') or 0) > 0:
                ws.cell(rn, ci2).fill = PatternFill('solid', fgColor=C_ERR)
        except (ValueError, TypeError):
            pass
    # Delayed Count > 0 → yellow
    if 'Delayed Count' in headers:
        ci3 = headers.index('Delayed Count') + 1
        try:
            if int(r.get('Delayed Count') or 0) > 0:
                ws.cell(rn, ci3).fill = PatternFill('solid', fgColor=C_WARN)
        except (ValueError, TypeError):
            pass


# ── Public API ────────────────────────────────────────────────────────────────

def generate_excel(summary, details, routes, exceptions,
                   store_insights=None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, 'Daily Summary',    summary,    status_col='__none__',
                 extra_color_fn=_color_summary)
    _write_sheet(wb, 'Delivery Details', details,    status_col='Status',
                 extra_color_fn=_color_details)
    _write_sheet(wb, 'Route Summary',    routes,     status_col='Status',
                 extra_color_fn=_color_routes)
    _write_sheet(wb, 'Exceptions',       exceptions, status_col='__none__',
                 extra_color_fn=_color_exc)

    if store_insights:
        _write_sheet(wb, 'Store Insights', store_insights, status_col='__none__',
                     hdr_bg='2D6A4F',
                     extra_color_fn=_color_store_insights)

    # ── Chart data ────────────────────────────────────────────────────────────
    cdata = wb.create_sheet('_data')
    cdata.sheet_state = 'hidden'

    people = {}
    for r in summary:
        n = r.get('Name', '')
        people[n] = people.get(n, 0) + (r.get('Total Deliveries (POD)') or 0)
    cdata['A1'], cdata['B1'] = 'Person', 'Deliveries'
    for i, (n, v) in enumerate(people.items(), 2):
        cdata[f'A{i}'] = n
        cdata[f'B{i}'] = v

    # Efficiency per person
    eff_data = {}
    for r in summary:
        n = r.get('Name', '')
        e = r.get('Efficiency %')
        if e is not None:
            if n not in eff_data:
                eff_data[n] = []
            eff_data[n].append(e)
    cdata['D1'], cdata['E1'] = 'Person', 'Avg Efficiency %'
    for i, (n, vals) in enumerate(eff_data.items(), 2):
        cdata[f'D{i}'] = n
        cdata[f'E{i}'] = round(sum(vals) / len(vals), 1)

    status_cnt: dict = {}
    for r in details:
        st = r.get('Status', 'OK')
        status_cnt[st] = status_cnt.get(st, 0) + 1
    cdata['G1'], cdata['H1'] = 'Status', 'Count'
    for i, (k, v) in enumerate(status_cnt.items(), 2):
        cdata[f'G{i}'] = k
        cdata[f'H{i}'] = v

    if people:
        cws = wb.create_sheet('Charts')

        c1 = BarChart()
        c1.type = 'col'
        c1.title = 'Deliveries per Person'
        c1.add_data(Reference(cdata, min_col=2, min_row=1, max_row=len(people)+1),
                    titles_from_data=True)
        c1.set_categories(Reference(cdata, min_col=1, min_row=2, max_row=len(people)+1))
        c1.width, c1.height = 16, 10
        cws.add_chart(c1, 'A1')

        c2 = BarChart()
        c2.type = 'col'
        c2.title = 'Delivery Status Breakdown'
        c2.add_data(Reference(cdata, min_col=8, min_row=1, max_row=len(status_cnt)+1),
                    titles_from_data=True)
        c2.set_categories(Reference(cdata, min_col=7, min_row=2, max_row=len(status_cnt)+1))
        c2.width, c2.height = 16, 10
        cws.add_chart(c2, 'A18')

        if eff_data:
            c3 = BarChart()
            c3.type = 'col'
            c3.title = 'Avg Efficiency % per Person'
            c3.add_data(Reference(cdata, min_col=5, min_row=1, max_row=len(eff_data)+1),
                        titles_from_data=True)
            c3.set_categories(Reference(cdata, min_col=4, min_row=2, max_row=len(eff_data)+1))
            c3.width, c3.height = 16, 10
            cws.add_chart(c3, 'J1')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
